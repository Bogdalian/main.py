import datetime
from pprint import pprint
from collections import defaultdict
from dateutil.parser import parse
from openpyxl import load_workbook
from connect_DB.connect import conn
from classes.classes_Get_Data_with_QUERY import Get_data_addendum
from classes.classes_Get_Data_with_QUERY import (QueryData_indicator_4,
                                                 QueryData_indicator_2_and_3)
from Query.queries   import (QUERY_4, QUERY_2_1,QUERY_2_2,QUERY_3_1,QUERY_3_2,
                             QUERY_1_1_add,QUERY_1_2_add,QUERY_1_3_add,QUERY_5)


def get_DataFrame_DB(year, num_week):
    all_dict = defaultdict(list)
    # ------------------------------------------------------------------------------------------------------------------
    b = QueryData_indicator_2_and_3(QUERY_2_1, QUERY_2_2, conn, year, num_week,1)
    full, part, date = b.get_raw_indicator_2()
    diff_2 = round(part / full * 100, 1)
    all_dict['98.18.1.1.1'].append(full)
    all_dict['98.18.1.1.2'].append(part)
    all_dict['98.18.1.1.3'].append(diff_2)
    # ------------------------------------------------------------------------------------------------------------------
    d = QueryData_indicator_4(QUERY_4, conn, year, num_week)
    data, data_esia = d.get_raw_indicator_4()
    diff_4 =  round((data_esia/data)*100, 1)
    all_dict['98.18.1.2.1'].append(data)
    all_dict['98.18.1.2.2'].append(data_esia)
    all_dict['98.18.1.2.3'].append(diff_4)
    all_dict['неделя'].append((f'{num_week} нед. {year}', 'Код')) # дата загрузочного

    # #№№№№№№№№№№№№№№№№№№№№№№№№№№№№№№№№№№№№№№№-------ПЕРВИЧНЫЕ---№№№№№№№№№№№№№№№№№№№№№№№№№№№№№№№№№№№№№№№№№№№№№№№№№№№№
    c = QueryData_indicator_2_and_3(QUERY_3_1, QUERY_3_2, conn, year, num_week, 1)
    full_, part, date = c.get_raw_indicator_2()
    date = date.split(' ')[1]
    date = parse(date)
    d = date.day
    m = date.month
    y = date.year
    date = f'{d}.{m}.{y}'
    date = str(datetime.datetime.strptime(date, '%d.%m.%Y').date())
    date = datetime.datetime.strptime(date, '%Y-%m-%d').strftime('%d.%m.%Y')
    app = Get_data_addendum(query1_1=QUERY_1_1_add,
                            query_1_2=QUERY_1_2_add,
                            query_1_3=QUERY_1_3_add,
                            num_week=num_week,
                            year= year,
                            number=1)
    full_data, rejected, accepted, precent_reject = app.get_data_query_1(app.start_date, app.end_date)
    all_dict['98.18.2.1'].append(full_data)
    all_dict['98.18.2.2'].append(rejected)
    all_dict['98.18.2.3'].append(precent_reject)
    all_dict['98.18.2.4'].append(accepted)
    mean_time_accept, mean_time_reject = app.get_mean_time_accept_and_reject()
    all_dict['98.18.4'].append(mean_time_accept)
    all_dict['98.18.5'].append(mean_time_reject)
    delta_reject_precent, full_data_precent = app.get_data_by_previous_week()
    all_dict['98.18.2.5'].append(delta_reject_precent)
    all_dict['98.18.2.6'].append(full_data_precent)
    # добавляем среднее время для от создания до ПРИНЯТИЯ И ОТКЛОНЕНИЯ
    app = Get_data_addendum(query_5=QUERY_5,
                            num_week=num_week,
                            year=year,
                            number=1)
    #mean_time_all = app.get_mean_time_accept_and_reject_ALL()
    all_dict['98.18.3'].append(round((mean_time_accept + mean_time_reject)/2))
    # ---------------------------------




    #----------------------------------
    ###################################----------------ВТОРИЧНЫЕ--------------№№№№№№№№№№№№№№№№№№№№#####№№№№№№№№№№№№№№№№
    app = Get_data_addendum(query1_1=QUERY_1_1_add,
                            query_1_2=QUERY_1_2_add,
                            query_1_3=QUERY_1_3_add,
                            num_week=num_week,
                            year=year,
                            number=2)
    full_data, rejected, accepted, precent_reject = app.get_data_query_1(app.start_date, app.end_date)
    all_dict['98.18.6.1'].append(full_data)
    all_dict['98.18.6.2'].append(rejected)
    all_dict['98.18.6.3'].append(precent_reject)
    all_dict['98.18.6.4'].append(accepted)
    mean_time_accept, mean_time_reject = app.get_mean_time_accept_and_reject()
    all_dict['98.18.6.8'].append(mean_time_accept)
    all_dict['98.18.6.9'].append(mean_time_reject)
    delta_reject_precent, full_data_precent = app.get_data_by_previous_week()
    all_dict['98.18.6.5'].append(delta_reject_precent)
    all_dict['98.18.6.6'].append(full_data_precent)
    #..................................................................
    # добавляем среднее время для от создания до ПРИНЯТИЯ И ОТКЛОНЕНИЯ
    app = Get_data_addendum(query_5=QUERY_5,
                            num_week=num_week,
                            year=year,
                            number=2)
    #mean_time_all = app.get_mean_time_accept_and_reject_ALL()
    all_dict['98.18.6.7'].append(round((mean_time_accept + mean_time_reject)/2))

    ######################################----------------ТРЕТИЧНЫЕ--------------#######################################
    app = Get_data_addendum(query1_1=QUERY_1_1_add,
                            query_1_2=QUERY_1_2_add,
                            query_1_3=QUERY_1_3_add,
                            num_week=num_week,
                            year=year,
                            number=3)
    full_data, rejected, accepted, precent_reject = app.get_data_query_1(app.start_date, app.end_date)
    all_dict['98.18.7.1'].append(full_data)
    all_dict['98.18.7.2'].append(rejected)
    all_dict['98.18.7.3'].append(precent_reject)
    all_dict['98.18.7.4'].append(accepted)
    mean_time_accept, mean_time_reject = app.get_mean_time_accept_and_reject()
    all_dict['98.18.7.8'].append(mean_time_accept)
    all_dict['98.18.7.9'].append(mean_time_reject)
    delta_reject_precent, full_data_precent = app.get_data_by_previous_week()
    all_dict['98.18.7.5'].append(delta_reject_precent)
    all_dict['98.18.7.6'].append(full_data_precent)
    # добавляем среднее время для от создания до ПРИНЯТИЯ И ОТКЛОНЕНИЯ
    app = Get_data_addendum(query_5=QUERY_5,
                            num_week=num_week,
                            year=year,
                            number=3)
    #mean_time_all = app.get_mean_time_accept_and_reject_ALL()
    all_dict['98.18.7.7'].append(round((mean_time_accept + mean_time_reject)/2))

    return all_dict, num_week, year, date
def get_data_for_load(year, num_week):
    data_dict, num_week, year, date = get_DataFrame_DB(year, num_week)
    wb = load_workbook(r'classes/test_template.xlsx')
    ws = wb.active
    ws['E1'].value = f'{num_week} нед. {year}'
    for i in range(1,43,1):
        try:
            ws[f'E{i}'].value = data_dict[ws[f'A{i}'].value][0]
        except (KeyError, IndexError):
            continue
    title = data_dict["неделя"][0][0].split('по')[0].replace('с', '').strip()
    title_file = f'Загрузочный {num_week} неделя {year} (c {date}).xlsx'
    wb.save(f"Z:\Управление ИАС и СЦ\Проекты вне ГК\Модераторы Наш СПб\Иерархии\Загрузочные\неделя\ {title_file}")
    # wb.save(f"C:\\Users\\b.bulatov\PycharmProjects\DataBase_Our_SPB_connect\ {title_file}")
    return title_file